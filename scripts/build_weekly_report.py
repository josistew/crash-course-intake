"""
build_weekly_report.py — Generates the Rez Weekly Report workbook with all formulas.

Produces: Rez-Weekly-Report.xlsx (saved in the project root)

Usage:
    cd /Users/josi/crash-course-intake
    python3 scripts/build_weekly_report.py [--output path/to/file.xlsx]

IMPORTANT: Column headers in import tabs are PLACEHOLDERS. After Rez's first
CSV export from each platform, update header rows in each import tab to match
actual column names. Formulas use MATCH() on these headers — they must match
exactly.
"""

import argparse
import os
import sys

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.comments import Comment
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip3 install openpyxl")
    sys.exit(1)

# Add scripts dir to path so sample_data imports correctly
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

from sample_data import (
    SQUARE_HEADERS, DOORDASH_HEADERS, UBEREATS_HEADERS,
    GRUBHUB_HEADERS, BEK_HEADERS, SQUARE_LABOR_HEADERS,
    LOCATIONS, BRANDS,
    SQUARE_SAMPLE_ROWS, DOORDASH_SAMPLE_ROWS, UBEREATS_SAMPLE_ROWS,
    GRUBHUB_SAMPLE_ROWS, BEK_SAMPLE_ROWS, SQUARE_LABOR_SAMPLE_ROWS,
    EMPLOYEE_ROSTER, PRIOR_WEEK_HEADERS, PRIOR_WEEK_SAMPLE,
    PAYROLL_SAMPLE_EMPLOYEES,
)

# ==============================================================================
# OUTPUT PATH
# ==============================================================================

PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
DEFAULT_OUTPUT = os.path.join(PROJECT_ROOT, "Rez-Weekly-Report.xlsx")

# ==============================================================================
# STYLE CONSTANTS
# ==============================================================================

# Tab colors
TAB_IMPORT  = "4A86C8"   # steel blue — import tabs
TAB_CALC    = "999999"   # neutral gray — location calc tabs
TAB_AMBER   = "F4A442"   # amber — Summary + Prior-Week (tabs Rez actually uses)
TAB_GREEN   = "6AA84F"   # sage green — Instructions + Employee-Roster (reference)

# Header background colors
HDR_IMPORT  = "4472C4"   # blue — import tab headers
HDR_CALC    = "2F5496"   # darker blue — calc/summary headers
HDR_INSTR   = "4A86C8"   # same as tab color for Instructions

# Fill colors
FILL_IMPORT_HDR  = PatternFill("solid", fgColor="4472C4")
FILL_CALC_HDR    = PatternFill("solid", fgColor="2F5496")
FILL_INSTR_HDR   = PatternFill("solid", fgColor="2F5496")
FILL_ALT_ROW     = PatternFill("solid", fgColor="F2F2F2")
FILL_WHITE       = PatternFill("solid", fgColor="FFFFFF")
FILL_VALID_BAD   = PatternFill("solid", fgColor="FBE4E4")  # light red — default validation
FILL_VALID_GOOD  = PatternFill("solid", fgColor="D9EAD3")  # light green — OK validation
FILL_MOTO_MEDI   = PatternFill("solid", fgColor="FDE9D9")  # light orange — MM brand band
FILL_TIKKA_SHACK = PatternFill("solid", fgColor="D0E2F1")  # light teal — TS brand band
FILL_AMBER_LIGHT = PatternFill("solid", fgColor="FFF2CC")  # light amber — summary title
FILL_SUMMARY_HDR = PatternFill("solid", fgColor="434343")  # dark — summary column headers
FILL_SECTION_HDR = PatternFill("solid", fgColor="E8F0FE")  # light blue — section headers
FILL_MANUAL_ENTRY = PatternFill("solid", fgColor="FFF9C4") # light yellow — manual entry cells

# Conditional formatting fills
FILL_CF_GREEN = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
FILL_CF_RED   = PatternFill(start_color="FBE4E4", end_color="FBE4E4", fill_type="solid")
FILL_CF_AMBER = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")  # milestone approaching (<=30 days)

# Fonts
FONT_DEFAULT = Font(name="Calibri", size=11)
FONT_HEADER  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BOLD    = Font(name="Calibri", size=11, bold=True)
FONT_TITLE   = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
FONT_SUMMARY_TITLE = Font(name="Calibri", size=18, bold=True)
FONT_VALID_BAD = Font(name="Calibri", size=10, color="CC0000")
FONT_VALID_OK  = Font(name="Calibri", size=10, color="006100")
FONT_SECTION   = Font(name="Calibri", size=12, bold=True, color="2F5496")
FONT_CF_GREEN  = Font(color="006100", bold=True, name="Calibri", size=11)
FONT_CF_RED    = Font(color="9C0006", bold=True, name="Calibri", size=11)

# Alignment
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_LEFT_NW = Alignment(horizontal="left",  vertical="top",    wrap_text=True)


# ==============================================================================
# LOCATION CONFIG
# Prior-Week tab has 5 KPIs per location. Column offsets (1-based within B:Z):
#   offset 1 = Net Revenue, 2 = Purchase Cost %, 3 = Labor Cost %,
#   offset 4 = Orders, 5 = Avg Ticket
# Location order: MML=1, MMA=2, MM3=3, TS1=4, TS2=5
# Column layout: A=Week Ending, B=MML Net Rev (col 2), ..., F=MML Avg Ticket (col 6),
#                G=MMA Net Rev (col 7), ..., Z=TS2 Avg Ticket (col 26)
# ==============================================================================

LOC_ORDER = ["MML", "MMA", "MM3", "TS1", "TS2"]

# For each location: full name as it appears in CSVs
LOC_FULL_NAMES = {
    "MML": "Moto Medi Lubbock",
    "MMA": "Moto Medi Amarillo",
    "MM3": "Moto Medi 3rd",
    "TS1": "Tikka Shack 1",
    "TS2": "Tikka Shack 2",
}

# Prior-Week column offsets for MATCH lookup
# Col A = Week Ending. Col B onward = KPIs in blocks of 5 per location.
# For location index i (0-based): Net Rev = 2 + 5*i, ..., Avg Ticket = 6 + 5*i
def prior_week_col_offset(loc_idx, metric_idx):
    """
    loc_idx: 0-based (MML=0, MMA=1, MM3=2, TS1=3, TS2=4)
    metric_idx: 0=Net Revenue, 1=Purchase Cost %, 2=Labor Cost %, 3=Orders, 4=Avg Ticket
    Returns 1-based column number in Prior-Week tab.
    """
    return 2 + (loc_idx * 5) + metric_idx


def col_width_for(text, padding=4):
    """Return a reasonable column width for the given text plus padding."""
    return max(len(str(text)) + padding, 10)


def style_header_cell(cell, text, fill=None, font=None, alignment=None):
    """Apply bold header styling to a cell."""
    cell.value = text
    cell.font  = font or FONT_HEADER
    cell.fill  = fill or FILL_IMPORT_HDR
    cell.alignment = alignment or ALIGN_CENTER


def apply_tab_color(ws, hex_color):
    ws.sheet_properties.tabColor = hex_color


def freeze(ws, cell_ref):
    ws.freeze_panes = cell_ref


# ==============================================================================
# WORKSHEET BUILDERS
# ==============================================================================

def build_instructions(ws):
    """Build the Instructions tab."""
    apply_tab_color(ws, TAB_GREEN)

    # Title row
    ws.merge_cells("A1:B1")
    cell = ws["A1"]
    cell.value     = "Rez Weekly Report — Instructions"
    cell.font      = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    cell.fill      = FILL_CALC_HDR
    cell.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 30

    # Column A width for readability
    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 20

    content = [
        ("", False),
        ("IMPORTANT: Column headers in each import tab are PLACEHOLDERS. After your first "
         "CSV export from each platform, update the header row in each import tab to match "
         "your actual column names. The formulas use MATCH() on these headers — they must "
         "match exactly.", False),
        ("", False),
        ("LABOR COST NOTE: The 'Est. Labor Cost (manual)' cell in each Location Calc tab "
         "requires manual entry for v1. Enter the total labor cost for that location from your "
         "Square Labor report each week. Labor Cost % is then computed automatically. Phase 2 "
         "will automate this via Square Labor CSV import.", False),
        ("", False),
        ("WEEKLY CADENCE", True),
        ("Every Monday morning: clear last week's data from each import tab (delete rows 3 "
         "and below — keep rows 1 and 2), paste new CSVs, then check the Summary tab.", False),
        ("", False),
        ("HOW TO EXPORT & PASTE — SQUARE", True),
        ("1. Log into Square Dashboard", False),
        ("2. Go to: Reports > Sales Summary", False),
        ("3. Select date range: Mon–Sun of the week you are reporting", False),
        ("4. Click 'Export CSV'", False),
        ("5. Open the CSV, select all, copy", False),
        ("6. Click into cell A3 of the Square-Import tab in this workbook", False),
        ("7. Paste (do not paste with formatting — plain paste only)", False),
        ("", False),
        ("HOW TO EXPORT & PASTE — DOORDASH", True),
        ("1. Log into DoorDash Merchant Portal", False),
        ("2. Go to: Financials > Report Builder", False),
        ("3. Select the week you are reporting (Monday to Sunday)", False),
        ("4. Download CSV", False),
        ("5. Open the CSV, select all, copy", False),
        ("6. Click into cell A3 of the DoorDash-Import tab", False),
        ("7. Paste (plain paste only)", False),
        ("", False),
        ("HOW TO EXPORT & PASTE — UBER EATS", True),
        ("1. Log into Uber Eats Manager", False),
        ("2. Go to: Analytics > Reports", False),
        ("3. Select 'Download Report' and choose the reporting week", False),
        ("4. Open the CSV, select all, copy", False),
        ("5. Click into cell A3 of the UberEats-Import tab", False),
        ("6. Paste (plain paste only)", False),
        ("", False),
        ("HOW TO EXPORT & PASTE — GRUBHUB", True),
        ("1. Log into Grubhub Restaurant Hub", False),
        ("2. Go to: Reports", False),
        ("3. Select the week and click 'Export'", False),
        ("4. Open the CSV, select all, copy", False),
        ("5. Click into cell A3 of the Grubhub-Import tab", False),
        ("6. Paste (plain paste only)", False),
        ("NOTE: Grubhub payouts are sometimes delayed 7–10 days. If the Grubhub column "
         "shows $0 on Monday, that is normal — check back Tuesday or Wednesday and update "
         "the Last Updated cell in the Grubhub-Import tab.", False),
        ("", False),
        ("HOW TO EXPORT & PASTE — BEK ENTREE", True),
        ("1. Log into BEK Entree Portal", False),
        ("2. Go to: Order History > Export Invoice CSV", False),
        ("3. Select date range matching your reporting week", False),
        ("4. Open the CSV, select all, copy", False),
        ("5. Click into cell A3 of the BEK-Import tab", False),
        ("6. Paste (plain paste only)", False),
        ("NOTE: BEK purchases are company-wide (not per-location). The workbook divides "
         "BEK total evenly across 5 locations as an estimate. For exact allocation, enter "
         "the per-location amount manually in the BEK Purchases cell of each Calc tab.", False),
        ("", False),
        ("LABOR COST (MANUAL ENTRY — v1)", True),
        ("After pasting all CSV data: open each Location Calc tab and enter total labor cost "
         "for that location in the yellow 'Est. Labor Cost (manual)' cell. Get this number from "
         "your Square Labor report (Total Hours x Average Hourly Rate for that location). "
         "Labor Cost % is computed automatically once you enter the dollar amount.", False),
        ("", False),
        ("TROUBLESHOOTING", True),
        ("If you see $0 or ERROR in the Summary tab, check the validation row (row 2) in "
         "each import tab. A red cell in row 2 means that column did not paste as a number "
         "— the data came in as text. Try pasting again without formatting, or paste into "
         "Notepad first and re-copy to strip formatting.", False),
        ("", False),
        ("If MATCH column errors (#N/A) appear, the column header in the import tab does not "
         "match the header in your CSV export. Update row 1 of the import tab to match the "
         "exact column name in your CSV file.", False),
    ]

    for r, (text, is_header) in enumerate(content, start=2):
        cell = ws.cell(row=r, column=1, value=text)
        if is_header:
            cell.font = FONT_SECTION
            cell.fill = PatternFill("solid", fgColor="E8F0FE")
            ws.row_dimensions[r].height = 20
        else:
            cell.font = FONT_DEFAULT
            cell.fill = FILL_WHITE
            ws.row_dimensions[r].height = 40 if len(text) > 80 else 18
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    freeze(ws, "A2")


# Key: list of column indices (0-based) that should be numeric in each import tab
NUMERIC_COLS = {
    "Square-Import":    {"Net Sales": 3, "Gross Sales": 4},
    "DoorDash-Import":  {"Net Payout": 4, "Order Count": 5},
    "UberEats-Import":  {"Net Payout": 4, "Trips": 5},
    "Grubhub-Import":   {"Net Payout": 5, "Orders": 6},
    "BEK-Import":       {"Total": 6, "Quantity": 4},
}


def build_import_tab(ws, tab_name, headers, sample_rows):
    """Build a platform import tab with headers, validation row, and sample data."""
    apply_tab_color(ws, TAB_IMPORT)

    # --- Row 1: PLACEHOLDER notice as a merged note above headers ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    notice_cell = ws.cell(row=1, column=1,
        value="PLACEHOLDER HEADERS — Update row 2 to match your actual CSV export column names exactly.")
    notice_cell.font  = Font(name="Calibri", size=10, bold=True, color="7F4F00")
    notice_cell.fill  = PatternFill("solid", fgColor="FFF2CC")
    notice_cell.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 20

    # --- Row 2: Column headers ---
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        style_header_cell(cell, header, fill=FILL_IMPORT_HDR)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width_for(header)

    freeze(ws, "A3")

    # --- Row 3: Validation row ---
    numeric_map = NUMERIC_COLS.get(tab_name, {})
    # Build a reverse map: header_name -> 1-based col index (based on position in headers list)
    header_col_map = {h: i for i, h in enumerate(headers, start=1)}

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        col_letter = get_column_letter(col_idx)

        if header in numeric_map or any(h in header for h in ["Net", "Payout", "Total", "Sales", "Count", "Orders", "Trips", "Quantity"]):
            # MATCH-based validation formula referencing row 4 (first data row)
            # Checks: does the value at (first_data_row, this_col) parse as a number?
            formula = (
                f'=IF(ISNUMBER(INDEX({tab_name}!A:Z,4,'
                f'MATCH("{header}",{tab_name}!2:2,0))),"OK",'
                f'"PASTE ERROR - {header} is not a number")'
            )
            cell.value = formula
            cell.fill  = FILL_VALID_BAD
            cell.font  = FONT_VALID_BAD
            cell.alignment = ALIGN_CENTER
        else:
            # Text/date columns — just note expected type
            cell.value = f'[{header} — text/date]'
            cell.fill  = PatternFill("solid", fgColor="F3F3F3")
            cell.font  = Font(name="Calibri", size=9, color="888888", italic=True)
            cell.alignment = ALIGN_CENTER

        ws.row_dimensions[3].height = 22

    # Conditional formatting: if cell = "OK" → green fill
    ok_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    ok_font = Font(color="006100", bold=True, name="Calibri", size=10)
    # Apply per-cell since row content varies
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.conditional_formatting.add(
            f"{col_letter}3",
            FormulaRule(
                formula=[f'={col_letter}3="OK"'],
                fill=ok_fill,
                font=ok_font,
                stopIfTrue=False,
            )
        )

    # --- Rows 4+: Sample data ---
    for r_idx, row_data in enumerate(sample_rows, start=4):
        for c_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = FONT_DEFAULT
            cell.alignment = ALIGN_LEFT
            # Alternating row color
            if r_idx % 2 == 0:
                cell.fill = FILL_ALT_ROW
            else:
                cell.fill = FILL_WHITE

    # --- Last Updated marker ---
    last_data_row = 3 + len(sample_rows)
    label_row = last_data_row + 2
    ws.cell(row=label_row, column=1, value="Last Updated:").font = FONT_BOLD
    date_cell = ws.cell(row=label_row, column=2, value="")
    date_cell.number_format = "YYYY-MM-DD"
    date_cell.fill = PatternFill("solid", fgColor="FFF9C4")
    comment = Comment("Enter the date you pasted this data.", "Build Script")
    date_cell.comment = comment


def build_labor_import(ws):
    """
    Build the Labor-Import tab for Square Labor CSV paste-in.

    Layout:
      Row 1: PLACEHOLDER notice (merged) with cell comment on A1
      Row 2: Column headers from SQUARE_LABOR_HEADERS with comment on B2
      Row 3: Validation row — ISNUMBER() for hours columns, ISTEXT() for name/location,
             type-label for date/time columns
      Row 4+: Sample data from SQUARE_LABOR_SAMPLE_ROWS
      Freeze pane at A4 so headers stay visible while scrolling data.

    NOTE: Headers are PLACEHOLDERS — must be updated to match Rez's actual
    Square Labor CSV export column names before any SUMIFS formulas are written.
    The ISNUMBER validation in row 3 uses MATCH-based references (same pattern
    as build_import_tab) so they survive column reordering in the real CSV.
    """
    apply_tab_color(ws, TAB_IMPORT)

    tab_name = "Labor-Import"
    headers  = SQUARE_LABOR_HEADERS  # ["Date","Employee Name","Location","Clock In","Clock Out",
                                      #  "Regular Hours","Overtime Hours","Total Hours"]

    # --- Row 1: PLACEHOLDER notice (merged across all columns) ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    notice_cell = ws.cell(row=1, column=1,
        value="PLACEHOLDER HEADERS — Update row 2 to match your actual Square Labor CSV export column names exactly.")
    notice_cell.font      = Font(name="Calibri", size=10, bold=True, color="7F4F00")
    notice_cell.fill      = PatternFill("solid", fgColor="FFF2CC")
    notice_cell.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 20

    # Cell comment on A1 — extra context for Rez
    a1_comment = Comment(
        "PLACEHOLDER — update headers to match Rez's actual Square Labor export column names. "
        "After your first export, paste the real headers into row 2 of this tab so SUMIFS "
        "formulas in the Overtime-Tracker and Payroll-Output tabs resolve correctly.",
        "Build Script"
    )
    ws.cell(row=1, column=1).comment = a1_comment

    # --- Row 2: Column headers ---
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        style_header_cell(cell, header, fill=FILL_IMPORT_HDR)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width_for(header)

    # Cell comment on B2 — Employee Name column used by Overtime-Tracker SUMIFS
    b2_comment = Comment(
        "This is the Employee Name column — used by Overtime-Tracker SUMIFS. "
        "Values must match the 'Square Name' column in Employee-Roster exactly "
        "(including capitalization and spacing).",
        "Build Script"
    )
    ws.cell(row=2, column=2).comment = b2_comment

    freeze(ws, "A4")

    # --- Row 3: Validation row ---
    # Columns by type:
    #   Date (col 0)      : type label
    #   Employee Name (1) : ISTEXT validation via MATCH
    #   Location (2)      : ISTEXT validation via MATCH
    #   Clock In (3)      : type label (time — text in most CSV exports)
    #   Clock Out (4)     : type label (time — text in most CSV exports)
    #   Regular Hours (5) : ISNUMBER validation via MATCH
    #   Overtime Hours (6): ISNUMBER validation via MATCH
    #   Total Hours (7)   : ISNUMBER validation via MATCH

    ISTEXT_COLS  = {"Employee Name", "Location"}
    ISNUMBER_COLS = {"Regular Hours", "Overtime Hours", "Total Hours"}

    for col_idx, header in enumerate(headers, start=1):
        cell       = ws.cell(row=3, column=col_idx)
        col_letter = get_column_letter(col_idx)

        if header in ISNUMBER_COLS:
            # MATCH-based formula — checks first data row (row 4) at this column's position
            formula = (
                f'=IF(ISNUMBER(INDEX({tab_name}!A:Z,4,'
                f'MATCH("{header}",{tab_name}!2:2,0))),"OK",'
                f'"PASTE ERROR - {header} is not a number")'
            )
            cell.value     = formula
            cell.fill      = FILL_VALID_BAD
            cell.font      = FONT_VALID_BAD
            cell.alignment = ALIGN_CENTER

        elif header in ISTEXT_COLS:
            # ISTEXT check — Employee Name and Location should be text strings
            formula = (
                f'=IF(ISTEXT(INDEX({tab_name}!A:Z,4,'
                f'MATCH("{header}",{tab_name}!2:2,0))),"OK",'
                f'"CHECK - {header} should be text")'
            )
            cell.value     = formula
            cell.fill      = FILL_VALID_BAD
            cell.font      = FONT_VALID_BAD
            cell.alignment = ALIGN_CENTER

        else:
            # Date / Clock In / Clock Out — text/date label only
            cell.value     = f'[{header} — text/date]'
            cell.fill      = PatternFill("solid", fgColor="F3F3F3")
            cell.font      = Font(name="Calibri", size=9, color="888888", italic=True)
            cell.alignment = ALIGN_CENTER

        ws.row_dimensions[3].height = 22

    # Conditional formatting on row 3: cells that return "OK" turn green
    ok_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    ok_font = Font(color="006100", bold=True, name="Calibri", size=10)
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.conditional_formatting.add(
            f"{col_letter}3",
            FormulaRule(
                formula=[f'={col_letter}3="OK"'],
                fill=ok_fill,
                font=ok_font,
                stopIfTrue=False,
            )
        )

    # --- Rows 4+: Sample data from SQUARE_LABOR_SAMPLE_ROWS ---
    for r_idx, row_data in enumerate(SQUARE_LABOR_SAMPLE_ROWS, start=4):
        for c_idx, value in enumerate(row_data, start=1):
            cell           = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font      = FONT_DEFAULT
            cell.alignment = ALIGN_LEFT
            if r_idx % 2 == 0:
                cell.fill = FILL_ALT_ROW
            else:
                cell.fill = FILL_WHITE

    # --- Last Updated marker ---
    last_data_row = 3 + len(SQUARE_LABOR_SAMPLE_ROWS)
    label_row     = last_data_row + 2
    ws.cell(row=label_row, column=1, value="Last Updated:").font = FONT_BOLD
    date_cell = ws.cell(row=label_row, column=2, value="")
    date_cell.number_format = "YYYY-MM-DD"
    date_cell.fill          = PatternFill("solid", fgColor="FFF9C4")
    date_cell.comment       = Comment("Enter the date you pasted this data.", "Build Script")


def build_employee_roster(ws):
    """
    Build the Employee-Roster reference tab.

    Columns A-G: core employee data (existing)
    Columns H-K: milestone tracking (added in Plan 02-01)
      H: Next Milestone Date  (static date string from EMPLOYEE_ROSTER[7])
      I: Milestone Type       (static text from EMPLOYEE_ROSTER[8])
      J: Days Until Milestone (formula: =IFERROR(H{row}-TODAY(),"") — integer)
      K: Milestone Status     (formula: =IF(J{row}="","",IF(J{row}<=0,"OVERDUE",IF(J{row}<=30,"SOON","OK"))))

    Conditional formatting on J2:J100:
      Red  (FILL_CF_RED)  + bold red:   =AND(J2<>"",J2<=0)  — overdue
      Amber (FILL_CF_AMBER) + bold:      =AND(J2>0,J2<=30)  — approaching
    """
    apply_tab_color(ws, TAB_GREEN)

    headers = [
        "Employee Name", "Square Name", "Location", "Hourly Rate",
        "Hire Date", "Pay Tier", "Notes",
        "Next Milestone Date", "Milestone Type", "Days Until Milestone", "Milestone Status",
    ]

    # Row 1: headers (A-K)
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        style_header_cell(cell, h, fill=FILL_CALC_HDR)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width_for(h, 6)

    freeze(ws, "A2")

    # Data validation: Location dropdown on column C
    loc_ids = ",".join(LOCATIONS.keys())
    loc_dv = DataValidation(
        type="list",
        formula1=f'"{loc_ids}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Location",
        error=f"Please select one of: {loc_ids}"
    )
    ws.add_data_validation(loc_dv)

    # Populate sample employees (cols A-I static; J and K are formulas)
    for r_idx, emp in enumerate(EMPLOYEE_ROSTER, start=2):
        name, sq_name, loc, rate, hire, tier, notes, milestone_date, milestone_type, _j, _k = emp

        # Columns A-I: static data
        static_row = [name, sq_name, loc, rate, hire, tier, notes, milestone_date, milestone_type]
        for c_idx, val in enumerate(static_row, start=1):
            cell           = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font      = FONT_DEFAULT
            cell.alignment = ALIGN_LEFT
            if r_idx % 2 == 0:
                cell.fill = FILL_ALT_ROW
            else:
                cell.fill = FILL_WHITE

        # Hourly Rate (col D=4): currency format
        ws.cell(row=r_idx, column=4).number_format = '"$"#,##0.00'
        # Hire Date (col E=5): date format
        ws.cell(row=r_idx, column=5).number_format = "YYYY-MM-DD"
        # Next Milestone Date (col H=8): date format (blank cells show nothing)
        ws.cell(row=r_idx, column=8).number_format = "YYYY-MM-DD"
        # Location (col C=3): add to data validation range
        loc_dv.add(ws.cell(row=r_idx, column=3))

        # Column J (10): Days Until Milestone formula
        # =IFERROR(H{row}-TODAY(),"") — subtracts today's date from milestone date
        # Returns "" when H is blank (IFERROR catches #VALUE! on empty string)
        j_cell           = ws.cell(row=r_idx, column=10)
        j_cell.value     = f'=IFERROR(H{r_idx}-TODAY(),"")'
        j_cell.font      = FONT_DEFAULT
        j_cell.alignment = ALIGN_CENTER
        j_cell.number_format = '0'  # integer days
        if r_idx % 2 == 0:
            j_cell.fill = FILL_ALT_ROW
        else:
            j_cell.fill = FILL_WHITE

        # Column K (11): Milestone Status formula
        # =IF(J="","", IF(J<=0,"OVERDUE", IF(J<=30,"SOON","OK")))
        k_cell           = ws.cell(row=r_idx, column=11)
        k_cell.value     = f'=IF(J{r_idx}="","",IF(J{r_idx}<=0,"OVERDUE",IF(J{r_idx}<=30,"SOON","OK")))'
        k_cell.font      = FONT_DEFAULT
        k_cell.alignment = ALIGN_CENTER
        if r_idx % 2 == 0:
            k_cell.fill = FILL_ALT_ROW
        else:
            k_cell.fill = FILL_WHITE

    # --- Conditional formatting on J2:J100 ---
    # Red: overdue (days <= 0, cell not blank)
    ws.conditional_formatting.add(
        "J2:J100",
        FormulaRule(
            formula=["AND(J2<>\"\",J2<=0)"],
            fill=FILL_CF_RED,
            font=Font(color="9C0006", bold=True, name="Calibri", size=11),
            stopIfTrue=True,
        )
    )
    # Amber: approaching (1–30 days remaining)
    ws.conditional_formatting.add(
        "J2:J100",
        FormulaRule(
            formula=["AND(J2>0,J2<=30)"],
            fill=FILL_CF_AMBER,
            font=Font(color="7F4F00", bold=True, name="Calibri", size=11),
            stopIfTrue=False,
        )
    )

    # --- Note row — spans A through K (all 11 columns) ---
    note_row = len(EMPLOYEE_ROSTER) + 3
    note_cell = ws.cell(row=note_row, column=1,
        value="NOTE: 'Square Name' must match the employee name exactly as it appears in Square's clock-in export. "
              "This is the VLOOKUP key for labor cost calculations. Nicknames, capitalization, and spacing must match. "
              "Milestone columns H-K: enter Next Milestone Date (col H) and Milestone Type (col I) manually; "
              "Days Until Milestone (col J) and Status (col K) update automatically via formula.")
    note_cell.font      = Font(name="Calibri", size=10, italic=True, color="666666")
    note_cell.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=11)
    ws.row_dimensions[note_row].height = 52


def build_prior_week(ws):
    """Build the Prior-Week snapshot tab."""
    apply_tab_color(ws, TAB_AMBER)

    headers = PRIOR_WEEK_HEADERS

    # Row 1: headers
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        style_header_cell(cell, h, fill=FILL_CALC_HDR)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width_for(h, 4)
    ws.row_dimensions[1].height = 30

    # Column A = date format
    ws.column_dimensions["A"].width = 16

    freeze(ws, "B2")

    # Row 2: prior week sample data (week ending 2026-03-09)
    for c_idx, val in enumerate(PRIOR_WEEK_SAMPLE, start=1):
        cell = ws.cell(row=2, column=c_idx, value=val)
        cell.font = FONT_DEFAULT
        cell.fill = FILL_ALT_ROW
        cell.alignment = ALIGN_CENTER
        if c_idx == 1:
            cell.number_format = "YYYY-MM-DD"
        elif "%" in headers[c_idx - 1]:
            cell.number_format = "0.0%"
        elif "Net Revenue" in headers[c_idx - 1]:
            cell.number_format = '"$"#,##0.00'

    # Row 3: empty (current week — to be filled from Summary after each Monday paste)
    empty_label = ws.cell(row=3, column=1, value="2026-03-16")
    empty_label.number_format = "YYYY-MM-DD"
    empty_label.font = Font(name="Calibri", size=11, color="AAAAAA", italic=True)
    empty_label.fill = FILL_WHITE
    note = Comment(
        "This row will hold the current week's KPIs after Monday reporting is complete. "
        "Copy values from Summary tab (paste as VALUES only) so next week's WoW formulas can reference them.",
        "Build Script"
    )
    empty_label.comment = note


def build_location_calc(ws, loc_id):
    """
    Build a Location Calc tab with full MATCH-based SUMIFS formulas.

    Row layout (matches formula references throughout):
      1:  Location name header
      2:  Week Ending date cell (linked to Summary!B2)
      3:  Column headers (This Week / Prior Week / WoW Change / WoW %)
      4:  "Revenue Breakdown" section header
      5:  Square Net Sales
      6:  DoorDash Net Revenue
      7:  UberEats Net Revenue
      8:  Grubhub Net Revenue
      9:  Total Net Revenue
      10: "Cost Metrics" section header
      11: [blank spacer]
      12: BEK Purchases
      13: Purchase Cost %
      14: "Labor" section header
      15: [blank spacer]
      16: Est. Labor Cost (manual entry)
      17: Labor Cost %
      18: "Volume" section header
      19: [blank spacer]
      20: Total Orders
      21: Avg Ticket Size
      22: [notes/blank]
    """
    apply_tab_color(ws, TAB_CALC)
    full_name = LOC_FULL_NAMES[loc_id]

    # Determine this location's Prior-Week column offset (0-based index)
    loc_idx = LOC_ORDER.index(loc_id)

    # --- Row 1: Location name merged across A-E ---
    ws.merge_cells("A1:E1")
    header_cell = ws["A1"]
    header_cell.value     = full_name
    header_cell.font      = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_cell.fill      = FILL_CALC_HDR
    header_cell.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 28

    # --- Row 2: Week Ending — linked to Summary!B2 ---
    ws.cell(row=2, column=1, value="Week Ending:").font = FONT_BOLD
    ws.cell(row=2, column=1).alignment = ALIGN_LEFT
    date_cell = ws.cell(row=2, column=2)
    date_cell.value = "='Summary'!B2"
    date_cell.number_format = "YYYY-MM-DD"
    date_cell.fill = PatternFill("solid", fgColor="FFF9C4")

    # --- Row 3: Column headers ---
    col_headers = {2: "This Week", 3: "Prior Week", 4: "WoW Change", 5: "WoW %"}
    for col, label in col_headers.items():
        cell = ws.cell(row=3, column=col, value=label)
        cell.font      = FONT_HEADER
        cell.fill      = FILL_CALC_HDR
        cell.alignment = ALIGN_CENTER

    # Column widths
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12

    # ===========================================================================
    # SECTION: Revenue Breakdown (rows 4-9)
    # ===========================================================================

    # Row 4: Section header
    ws.merge_cells("A4:E4")
    section_cell = ws.cell(row=4, column=1, value="Revenue Breakdown")
    section_cell.font = FONT_SECTION
    section_cell.fill = FILL_SECTION_HDR

    # Helper: write a revenue formula row
    def write_revenue_row(row, label, import_tab, amount_col, loc_col, loc_match_col=None, is_total=False):
        """
        Write a metric row with This Week formula (col B), Prior Week (col C),
        WoW Change (col D), WoW % (col E).
        """
        # Label (col A)
        cell_a = ws.cell(row=row, column=1, value=label)
        cell_a.font = FONT_BOLD if is_total else FONT_DEFAULT
        cell_a.alignment = ALIGN_LEFT
        if is_total:
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = FILL_SECTION_HDR

        # This Week formula (col B)
        cell_b = ws.cell(row=row, column=2)
        if is_total:
            cell_b.value = "=SUM(B5:B8)"
        else:
            if loc_match_col:
                # SUMIFS with location filter
                cell_b.value = (
                    f"=IFERROR(SUMIFS("
                    f"INDEX('{import_tab}'!A:Z,0,MATCH(\"{amount_col}\",'{import_tab}'!1:1,0)),"
                    f"INDEX('{import_tab}'!A:Z,0,MATCH(\"{loc_match_col}\",'{import_tab}'!1:1,0)),"
                    f"\"{full_name}\""
                    f"),0)"
                )
            else:
                cell_b.value = f"=0"  # fallback

        # Prior Week (col C) — INDEX/MATCH on Prior-Week tab by date (B2-7)
        pw_col = prior_week_col_offset(loc_idx, 0)  # Net Revenue column for this loc
        cell_c = ws.cell(row=row, column=3)
        if is_total:
            # Total Net Revenue references the Prior-Week Net Revenue column
            cell_c.value = (
                f"=IFERROR(INDEX('Prior-Week'!A:Z,"
                f"MATCH(B2-7,'Prior-Week'!A:A,0),"
                f"{pw_col}),0)"
            )
        else:
            # Component rows (Square, DoorDash, UberEats, Grubhub) not stored individually
            # in Prior-Week tab — show 0 so WoW % formula evaluates cleanly
            cell_c.value = 0

        # WoW Change (col D)
        cell_d = ws.cell(row=row, column=4)
        if is_total:
            cell_d.value = "=B9-C9"
        else:
            # Component rows have Prior Week = 0 (not tracked individually), so WoW = B-0
            cell_d.value = f"=B{row}-C{row}"

        # WoW % (col E)
        cell_e = ws.cell(row=row, column=5)
        if is_total:
            cell_e.value = "=IF(C9>0,(B9-C9)/C9,0)"
        else:
            cell_e.value = f"=IF(C{row}>0,(B{row}-C{row})/C{row},0)"

        # Number formats
        for c in [2, 3, 4]:
            ws.cell(row=row, column=c).number_format = '"$"#,##0'
        ws.cell(row=row, column=5).number_format = '0.0%'

    # Row 5: Square Net Sales
    write_revenue_row(5, "Square Net Sales",    "Square-Import",   "Net Sales",  "Location", "Location")
    # Row 6: DoorDash
    write_revenue_row(6, "DoorDash Net Revenue", "DoorDash-Import", "Net Payout", "Store Name", "Store Name")
    # Row 7: UberEats
    write_revenue_row(7, "UberEats Net Revenue", "UberEats-Import", "Net Payout", "Restaurant Name", "Restaurant Name")
    # Row 8: Grubhub
    write_revenue_row(8, "Grubhub Net Revenue",  "Grubhub-Import",  "Net Payout", "Restaurant", "Restaurant")
    # Row 9: Total Net Revenue (is_total=True)
    write_revenue_row(9, "Total Net Revenue", None, None, None, None, is_total=True)

    # WoW conditional formatting on E9 (Total Net Revenue WoW %)
    ws.conditional_formatting.add(
        "E9",
        FormulaRule(formula=["E9>0.05"], fill=FILL_CF_GREEN, font=FONT_CF_GREEN)
    )
    ws.conditional_formatting.add(
        "E9",
        FormulaRule(formula=["E9<-0.05"], fill=FILL_CF_RED, font=FONT_CF_RED)
    )

    # ===========================================================================
    # SECTION: Cost Metrics (rows 10-13)
    # ===========================================================================

    # Row 10: Section header
    ws.merge_cells("A10:E10")
    ws.cell(row=10, column=1, value="Cost Metrics").font = FONT_SECTION
    ws.cell(row=10, column=1).fill = FILL_SECTION_HDR

    # Row 11: blank spacer
    ws.row_dimensions[11].height = 6

    # Row 12: BEK Purchases
    # BEK is company-wide — sum all Food rows, divide by 5 locations
    # If BEK has no "Category" column, MATCH returns error and we fall back to sum all items / 5
    bek_formula = (
        "=IFERROR("
        "SUMIF("
        "INDEX('BEK-Import'!A:Z,0,MATCH(\"Category\",'BEK-Import'!1:1,0)),"
        "\"Food\","
        "INDEX('BEK-Import'!A:Z,0,MATCH(\"Total\",'BEK-Import'!1:1,0))"
        ")/5,"
        "IFERROR("
        "SUMIF("
        "INDEX('BEK-Import'!A:Z,0,MATCH(\"Total\",'BEK-Import'!1:1,0)),"
        "\">0\","
        "INDEX('BEK-Import'!A:Z,0,MATCH(\"Total\",'BEK-Import'!1:1,0))"
        ")/5,"
        "0)"
        ")"
    )
    cell_b12 = ws.cell(row=12, column=1, value="BEK Purchases")
    cell_b12.font = FONT_DEFAULT
    cell_b12.alignment = ALIGN_LEFT

    ws.cell(row=12, column=2).value = bek_formula
    ws.cell(row=12, column=2).number_format = '"$"#,##0'

    # Prior Week for BEK purchases — not tracked in Prior-Week tab per KPI (no direct column)
    # Leave blank; BEK is an allocation estimate anyway
    ws.cell(row=12, column=3).value = ""
    ws.cell(row=12, column=4).value = ""
    ws.cell(row=12, column=5).value = ""

    # Add comment explaining BEK allocation
    bek_comment = Comment(
        "BEK purchases are company-wide (not per-location). This divides the total food "
        "purchases by 5 locations as an estimate. If BEK CSV has no 'Category' column, "
        "it sums all invoice line items divided by 5. For precise allocation, enter manually.",
        "Build Script"
    )
    ws.cell(row=12, column=2).comment = bek_comment

    # Row 13: Purchase Cost %
    pw_col_purchase = prior_week_col_offset(loc_idx, 1)  # Purchase Cost % column
    ws.cell(row=13, column=1, value="Purchase Cost %").font = FONT_DEFAULT
    ws.cell(row=13, column=1).alignment = ALIGN_LEFT
    ws.cell(row=13, column=2).value = "=IF(B9>0,B12/B9,0)"
    ws.cell(row=13, column=2).number_format = '0.0%'
    ws.cell(row=13, column=3).value = (
        f"=IFERROR(INDEX('Prior-Week'!A:Z,"
        f"MATCH(B2-7,'Prior-Week'!A:A,0),"
        f"{pw_col_purchase}),0)"
    )
    ws.cell(row=13, column=3).number_format = '0.0%'
    ws.cell(row=13, column=4).value = "=B13-C13"
    ws.cell(row=13, column=4).number_format = '0.0%'
    ws.cell(row=13, column=5).value = "=IF(C13>0,(B13-C13)/C13,0)"
    ws.cell(row=13, column=5).number_format = '0.0%'

    # Conditional formatting on E13
    ws.conditional_formatting.add(
        "E13",
        FormulaRule(formula=["E13>0.05"], fill=FILL_CF_GREEN, font=FONT_CF_GREEN)
    )
    ws.conditional_formatting.add(
        "E13",
        FormulaRule(formula=["E13<-0.05"], fill=FILL_CF_RED, font=FONT_CF_RED)
    )

    # ===========================================================================
    # SECTION: Labor (rows 14-17)
    # ===========================================================================

    # Row 14: Section header
    ws.merge_cells("A14:E14")
    ws.cell(row=14, column=1, value="Labor").font = FONT_SECTION
    ws.cell(row=14, column=1).fill = FILL_SECTION_HDR

    # Row 15: blank spacer
    ws.row_dimensions[15].height = 6

    # Row 16: Estimated Labor Cost — MANUAL ENTRY for v1
    ws.cell(row=16, column=1, value="Est. Labor Cost (manual)").font = FONT_DEFAULT
    ws.cell(row=16, column=1).alignment = ALIGN_LEFT
    # Manual entry cell — yellow background to indicate input required
    labor_cell = ws.cell(row=16, column=2, value=0)
    labor_cell.number_format = '"$"#,##0'
    labor_cell.fill = FILL_MANUAL_ENTRY
    labor_comment = Comment(
        "MANUAL ENTRY — v1: Enter total labor cost for this location from your Square "
        "Labor report (Total Hours x Avg Hourly Rate). Phase 2 will automate this via "
        "Square Labor CSV import. Example: if 80 hours @ avg $14.50/hr, enter 1160.",
        "Build Script"
    )
    labor_cell.comment = labor_comment

    # Prior Week labor — from Prior-Week tab
    pw_col_labor = prior_week_col_offset(loc_idx, 2)  # Labor Cost % column
    # Prior Week for labor $ — not stored (only %). Leave blank.
    ws.cell(row=16, column=3).value = ""
    ws.cell(row=16, column=4).value = ""
    ws.cell(row=16, column=5).value = ""

    # Row 17: Labor Cost %
    ws.cell(row=17, column=1, value="Labor Cost %").font = FONT_DEFAULT
    ws.cell(row=17, column=1).alignment = ALIGN_LEFT
    ws.cell(row=17, column=2).value = "=IF(B9>0,B16/B9,0)"
    ws.cell(row=17, column=2).number_format = '0.0%'
    ws.cell(row=17, column=3).value = (
        f"=IFERROR(INDEX('Prior-Week'!A:Z,"
        f"MATCH(B2-7,'Prior-Week'!A:A,0),"
        f"{pw_col_labor}),0)"
    )
    ws.cell(row=17, column=3).number_format = '0.0%'
    ws.cell(row=17, column=4).value = "=B17-C17"
    ws.cell(row=17, column=4).number_format = '0.0%'
    ws.cell(row=17, column=5).value = "=IF(C17>0,(B17-C17)/C17,0)"
    ws.cell(row=17, column=5).number_format = '0.0%'

    # Conditional formatting on E17
    ws.conditional_formatting.add(
        "E17",
        FormulaRule(formula=["E17>0.05"], fill=FILL_CF_GREEN, font=FONT_CF_GREEN)
    )
    ws.conditional_formatting.add(
        "E17",
        FormulaRule(formula=["E17<-0.05"], fill=FILL_CF_RED, font=FONT_CF_RED)
    )

    # ===========================================================================
    # SECTION: Volume (rows 18-21)
    # ===========================================================================

    # Row 18: Section header
    ws.merge_cells("A18:E18")
    ws.cell(row=18, column=1, value="Volume").font = FONT_SECTION
    ws.cell(row=18, column=1).fill = FILL_SECTION_HDR

    # Row 19: blank spacer
    ws.row_dimensions[19].height = 6

    # Row 20: Total Orders — sum across all platforms
    orders_formula = (
        f"=IFERROR(SUMIFS("
        f"INDEX('Square-Import'!A:Z,0,MATCH(\"Order Count\",'Square-Import'!1:1,0)),"
        f"INDEX('Square-Import'!A:Z,0,MATCH(\"Location\",'Square-Import'!1:1,0)),"
        f"\"{full_name}\""
        f"),0)"
        f"+IFERROR(SUMIFS("
        f"INDEX('DoorDash-Import'!A:Z,0,MATCH(\"Order Count\",'DoorDash-Import'!1:1,0)),"
        f"INDEX('DoorDash-Import'!A:Z,0,MATCH(\"Store Name\",'DoorDash-Import'!1:1,0)),"
        f"\"{full_name}\""
        f"),0)"
        f"+IFERROR(SUMIFS("
        f"INDEX('UberEats-Import'!A:Z,0,MATCH(\"Trips\",'UberEats-Import'!1:1,0)),"
        f"INDEX('UberEats-Import'!A:Z,0,MATCH(\"Restaurant Name\",'UberEats-Import'!1:1,0)),"
        f"\"{full_name}\""
        f"),0)"
        f"+IFERROR(SUMIFS("
        f"INDEX('Grubhub-Import'!A:Z,0,MATCH(\"Orders\",'Grubhub-Import'!1:1,0)),"
        f"INDEX('Grubhub-Import'!A:Z,0,MATCH(\"Restaurant\",'Grubhub-Import'!1:1,0)),"
        f"\"{full_name}\""
        f"),0)"
    )

    ws.cell(row=20, column=1, value="Total Orders").font = FONT_DEFAULT
    ws.cell(row=20, column=1).alignment = ALIGN_LEFT
    ws.cell(row=20, column=2).value = orders_formula
    ws.cell(row=20, column=2).number_format = '#,##0'

    # Prior Week orders
    pw_col_orders = prior_week_col_offset(loc_idx, 3)  # Orders column
    ws.cell(row=20, column=3).value = (
        f"=IFERROR(INDEX('Prior-Week'!A:Z,"
        f"MATCH(B2-7,'Prior-Week'!A:A,0),"
        f"{pw_col_orders}),0)"
    )
    ws.cell(row=20, column=3).number_format = '#,##0'
    ws.cell(row=20, column=4).value = "=B20-C20"
    ws.cell(row=20, column=4).number_format = '#,##0'
    ws.cell(row=20, column=5).value = "=IF(C20>0,(B20-C20)/C20,0)"
    ws.cell(row=20, column=5).number_format = '0.0%'

    # Conditional formatting on E20
    ws.conditional_formatting.add(
        "E20",
        FormulaRule(formula=["E20>0.05"], fill=FILL_CF_GREEN, font=FONT_CF_GREEN)
    )
    ws.conditional_formatting.add(
        "E20",
        FormulaRule(formula=["E20<-0.05"], fill=FILL_CF_RED, font=FONT_CF_RED)
    )

    # Row 21: Avg Ticket
    ws.cell(row=21, column=1, value="Avg Ticket Size").font = FONT_DEFAULT
    ws.cell(row=21, column=1).alignment = ALIGN_LEFT
    ws.cell(row=21, column=2).value = "=IF(B20>0,B9/B20,0)"
    ws.cell(row=21, column=2).number_format = '"$"#,##0.00'

    # Prior Week avg ticket
    pw_col_ticket = prior_week_col_offset(loc_idx, 4)  # Avg Ticket column
    ws.cell(row=21, column=3).value = (
        f"=IFERROR(INDEX('Prior-Week'!A:Z,"
        f"MATCH(B2-7,'Prior-Week'!A:A,0),"
        f"{pw_col_ticket}),0)"
    )
    ws.cell(row=21, column=3).number_format = '"$"#,##0.00'
    ws.cell(row=21, column=4).value = "=B21-C21"
    ws.cell(row=21, column=4).number_format = '"$"#,##0.00'
    ws.cell(row=21, column=5).value = "=IF(C21>0,(B21-C21)/C21,0)"
    ws.cell(row=21, column=5).number_format = '0.0%'

    # Conditional formatting on E21
    ws.conditional_formatting.add(
        "E21",
        FormulaRule(formula=["E21>0.05"], fill=FILL_CF_GREEN, font=FONT_CF_GREEN)
    )
    ws.conditional_formatting.add(
        "E21",
        FormulaRule(formula=["E21<-0.05"], fill=FILL_CF_RED, font=FONT_CF_RED)
    )

    # Row 22: thin separator, then notes
    note_cell = ws.cell(row=23, column=1,
        value=f"[{loc_id}] Labor Cost requires manual entry — enter weekly labor total in the yellow B16 cell. "
              "Phase 2 automates via Square Labor CSV import (PAY-01).")
    note_cell.font = Font(name="Calibri", size=10, italic=True, color="AAAAAA")
    note_cell.alignment = Alignment(wrap_text=True, horizontal="left")
    ws.merge_cells(f"A23:E23")
    ws.row_dimensions[23].height = 32

    freeze(ws, "B4")


def build_summary(ws):
    """
    Build the Summary tab with cross-tab references, WoW display, and conditional formatting.

    Row layout:
      1:  Title
      2:  Week Ending (master date cell — drives all WoW lookups via calc tabs)
      3:  blank separator
      4:  Location column headers (MML | MMA | MM3 | TS1 | TS2)
      5:  Net Revenue row
      6:  WoW Net Revenue row
      7:  Purchase Cost % row
      8:  WoW Purchase Cost % row
      9:  Labor Cost % row
      10: WoW Labor Cost % row
      11: Order Volume row
      12: WoW Order Volume row
      13: Avg Ticket row
      14: WoW Avg Ticket row
      15: blank
      16: blank
      17: Notes section header
      18: Note: Purchase Cost %
      19: Note: Labor Cost %
      20: Note: Delivery revenue
    """
    apply_tab_color(ws, TAB_AMBER)

    # --- Row 1: Title ---
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value     = "Weekly Report Summary"
    title_cell.font      = Font(name="Calibri", size=18, bold=True, color="1C3557")
    title_cell.fill      = FILL_AMBER_LIGHT
    title_cell.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 36

    # --- Row 2: Week Ending (master date — all calc tabs link to Summary!B2) ---
    ws.cell(row=2, column=1, value="Week Ending:").font = FONT_BOLD
    ws.cell(row=2, column=1).alignment = ALIGN_LEFT
    week_cell = ws.cell(row=2, column=2)
    week_cell.number_format = "YYYY-MM-DD"
    week_cell.fill = FILL_MANUAL_ENTRY
    week_note = Comment(
        "Enter the Sunday date for this reporting week (e.g., 2026-03-15). "
        "All calc tabs link to this cell, so WoW comparisons update automatically.",
        "Build Script"
    )
    week_cell.comment = week_note
    ws.row_dimensions[2].height = 20

    # --- Row 3: blank separator ---
    ws.row_dimensions[3].height = 8

    # --- Row 4: Column headers ---
    loc_order = LOC_ORDER  # ["MML", "MMA", "MM3", "TS1", "TS2"]
    loc_full  = [LOCATIONS[l] for l in loc_order]

    header_labels = [""] + loc_full
    for c_idx, label in enumerate(header_labels, start=1):
        cell = ws.cell(row=4, column=c_idx, value=label)
        cell.font      = FONT_HEADER
        cell.fill      = FILL_SUMMARY_HDR
        cell.alignment = ALIGN_CENTER
    ws.row_dimensions[4].height = 28

    # --- Brand color bands (rows 5-14) ---
    for row in range(5, 15):
        for col in range(2, 5):   # B-D = MML, MMA, MM3 = Moto Medi
            ws.cell(row=row, column=col).fill = FILL_MOTO_MEDI
        for col in range(5, 7):   # E-F = TS1, TS2 = Tikka Shack
            ws.cell(row=row, column=col).fill = FILL_TIKKA_SHACK

    # Column letters for each location
    loc_cols = {
        "MML": 2, "MMA": 3, "MM3": 4, "TS1": 5, "TS2": 6
    }

    # -------------------------------------------------------------------
    # KPI reference rows
    # Calc tab row references:
    #   Net Revenue   → {loc}-Calc!B9  (Prior: C9, WoW%: E9)
    #   Purchase Cost → {loc}-Calc!B13 (Prior: C13, WoW%: E13)
    #   Labor Cost    → {loc}-Calc!B17 (Prior: C17, WoW%: E17)
    #   Orders        → {loc}-Calc!B20 (Prior: C20, WoW%: E20)
    #   Avg Ticket    → {loc}-Calc!B21 (Prior: C21, WoW%: E21)
    # -------------------------------------------------------------------

    kpi_defs = [
        # (metric_row, wow_row, label, calc_row, fmt_value, fmt_wow)
        (5,  6,  "Net Revenue",     9,  '"$"#,##0',  '"+0.0%;-0.0%"'),
        (7,  8,  "Purchase Cost %", 13, '0.0%',      '"+0.0%;-0.0%"'),
        (9,  10, "Labor Cost %",    17, '0.0%',      '"+0.0%;-0.0%"'),
        (11, 12, "Order Volume",    20, '#,##0',     '"+0.0%;-0.0%"'),
        (13, 14, "Avg Ticket",      21, '"$"#,##0.00', '"+0.0%;-0.0%"'),
    ]

    for (metric_row, wow_row, label, calc_row, fmt_val, fmt_wow) in kpi_defs:
        # --- Metric row label (col A) ---
        cell_a = ws.cell(row=metric_row, column=1, value=label)
        cell_a.font = FONT_BOLD
        cell_a.alignment = ALIGN_LEFT

        # --- WoW row label (col A) ---
        cell_wow_a = ws.cell(row=wow_row, column=1, value="  WoW")
        cell_wow_a.font = Font(name="Calibri", size=9, italic=True, color="666666")
        cell_wow_a.alignment = ALIGN_LEFT
        ws.row_dimensions[wow_row].height = 16

        for loc_id in LOC_ORDER:
            col = loc_cols[loc_id]
            calc_tab = f"{loc_id}-Calc"

            # Metric value cell
            metric_cell = ws.cell(row=metric_row, column=col)
            metric_cell.value = f"='{calc_tab}'!B{calc_row}"
            metric_cell.number_format = fmt_val
            metric_cell.alignment = ALIGN_CENTER

            # WoW display cell — raw percentage (colored by conditional formatting)
            wow_cell = ws.cell(row=wow_row, column=col)
            wow_cell.value = f"='{calc_tab}'!E{calc_row}"
            wow_cell.number_format = '+0.0%;-0.0%'
            wow_cell.alignment = ALIGN_CENTER
            wow_cell.font = Font(name="Calibri", size=9)

            # Conditional formatting on WoW cell
            col_letter = get_column_letter(col)
            cf_ref = f"='{calc_tab}'!E{calc_row}"
            ws.conditional_formatting.add(
                f"{col_letter}{wow_row}",
                FormulaRule(
                    formula=[f"'{calc_tab}'!E{calc_row}>0.05"],
                    fill=FILL_CF_GREEN,
                    font=Font(name="Calibri", size=9, color="006100", bold=True)
                )
            )
            ws.conditional_formatting.add(
                f"{col_letter}{wow_row}",
                FormulaRule(
                    formula=[f"'{calc_tab}'!E{calc_row}<-0.05"],
                    fill=FILL_CF_RED,
                    font=Font(name="Calibri", size=9, color="9C0006", bold=True)
                )
            )

    # --- Row 15-16: blank separators ---
    ws.row_dimensions[15].height = 8
    ws.row_dimensions[16].height = 8

    # --- Row 17: Notes section header ---
    ws.merge_cells("A17:F17")
    notes_header = ws.cell(row=17, column=1, value="Notes")
    notes_header.font = FONT_SECTION
    notes_header.fill = FILL_SECTION_HDR

    # --- Rows 18-20: Note text ---
    notes = [
        "Purchase Cost % = BEK food purchases (company-wide, divided by 5 locations) / net revenue. "
        "Not true COGS — does not include inventory variance or waste.",
        "Labor Cost % is estimated from manual entry in each Location Calc tab. "
        "Phase 2 automates this via Square Labor CSV import.",
        "All delivery revenue (DoorDash, UberEats, Grubhub) uses net payout after platform fees/commissions.",
    ]
    for i, note_text in enumerate(notes, start=18):
        ws.merge_cells(f"A{18 + i - 18}:F{18 + i - 18}")
        cell = ws.cell(row=18 + i - 18, column=1, value=note_text)
        cell.font = Font(name="Calibri", size=10, italic=True, color="444444")
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[18 + i - 18].height = 30

    # Freeze row 4 and column A (freeze at B5)
    freeze(ws, "B5")

    # Column widths
    ws.column_dimensions["A"].width = 20
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 18

    # Print area
    ws.print_area = "A1:F20"

    # Add thin bottom borders between metric groups
    thin_side = Side(style="thin", color="CCCCCC")
    thin_border = Border(bottom=thin_side)
    for row_num in [6, 8, 10, 12, 14]:
        for col in range(1, 7):
            ws.cell(row=row_num, column=col).border = thin_border


# ==============================================================================
# MAIN
# ==============================================================================

def build_workbook(output_path=None):
    if output_path is None:
        output_path = DEFAULT_OUTPUT

    print("Building Rez Weekly Report workbook...")
    print(f"Output: {output_path}")

    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Tab definitions: (name, tab_color, builder_func)
    tab_defs = [
        ("Instructions",    TAB_GREEN,  lambda ws: build_instructions(ws)),
        ("Square-Import",   TAB_IMPORT, lambda ws: build_import_tab(ws, "Square-Import",    SQUARE_HEADERS,   SQUARE_SAMPLE_ROWS)),
        ("DoorDash-Import", TAB_IMPORT, lambda ws: build_import_tab(ws, "DoorDash-Import",  DOORDASH_HEADERS, DOORDASH_SAMPLE_ROWS)),
        ("UberEats-Import", TAB_IMPORT, lambda ws: build_import_tab(ws, "UberEats-Import",  UBEREATS_HEADERS, UBEREATS_SAMPLE_ROWS)),
        ("Grubhub-Import",  TAB_IMPORT, lambda ws: build_import_tab(ws, "Grubhub-Import",   GRUBHUB_HEADERS,  GRUBHUB_SAMPLE_ROWS)),
        ("BEK-Import",      TAB_IMPORT, lambda ws: build_import_tab(ws, "BEK-Import",       BEK_HEADERS,      BEK_SAMPLE_ROWS)),
        ("Employee-Roster", TAB_GREEN,  lambda ws: build_employee_roster(ws)),
        ("Prior-Week",      TAB_AMBER,  lambda ws: build_prior_week(ws)),
        ("MML-Calc",        TAB_CALC,   lambda ws: build_location_calc(ws, "MML")),
        ("MMA-Calc",        TAB_CALC,   lambda ws: build_location_calc(ws, "MMA")),
        ("MM3-Calc",        TAB_CALC,   lambda ws: build_location_calc(ws, "MM3")),
        ("TS1-Calc",        TAB_CALC,   lambda ws: build_location_calc(ws, "TS1")),
        ("TS2-Calc",        TAB_CALC,   lambda ws: build_location_calc(ws, "TS2")),
        ("Summary",         TAB_AMBER,  lambda ws: build_summary(ws)),
    ]

    for tab_name, tab_color, builder_fn in tab_defs:
        print(f"  Building tab: {tab_name}...")
        ws = wb.create_sheet(title=tab_name)
        apply_tab_color(ws, tab_color)
        builder_fn(ws)

    print(f"\nSaving workbook to: {output_path}")
    wb.save(output_path)
    print(f"Done. File saved: {output_path}")
    print(f"Total tabs: {len(tab_defs)}")

    # Quick sanity check
    wb2 = openpyxl.load_workbook(output_path)
    expected_tabs = [t[0] for t in tab_defs]
    actual_tabs   = wb2.sheetnames
    if actual_tabs != expected_tabs:
        print(f"WARNING: Tab order mismatch.\n  Expected: {expected_tabs}\n  Got:      {actual_tabs}")
    else:
        print(f"Verified: All {len(expected_tabs)} tabs present in correct order.")

    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build Rez Weekly Report workbook")
    parser.add_argument("--output", default=None,
                        help="Output path for the .xlsx file (default: Rez-Weekly-Report.xlsx in project root)")
    args = parser.parse_args()
    build_workbook(output_path=args.output)
